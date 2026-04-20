"""
AI Line Item Extractor — Uses Azure OpenAI GPT-4o to extract line items from documents.

Supports: PDF, images (PNG/JPG/TIFF), Excel, CSV, Word (.docx)
"""

import base64
import csv
import io
import json
import logging
import os

logger = logging.getLogger(__name__)

# GPT-4o pricing (USD)
COST_PER_1M_INPUT_TOKENS = 2.50
COST_PER_1M_OUTPUT_TOKENS = 10.00

MAX_ITEMS = 100

SYSTEM_PROMPT = """You are a precise document-analysis assistant that extracts purchasable line items from business documents.

## Your Task
Identify every distinct purchasable line item in the document and return structured data for each one.

## What Counts as a Line Item
A line item is any individually priced product, service, material, labour charge, fee, or deliverable listed in the document.  Common indicators:
- A row in a table with columns like #, Qty, Description, Unit Price, Total Price (or similar headings)
- A numbered or bulleted entry with a description and dollar amount
- An entry on a quote, invoice, purchase order, proposal, or estimate
- Items in a bill of materials, parts list, or cost breakdown

## Extraction Rules
1. **description** — The item name or description.  Use the text from the document verbatim; do NOT paraphrase.  If a part number, SKU, or catalog number is present, *prepend* it (e.g. "PT-4420 — Hydraulic pump").
2. **quantity** — Integer number of units.  Default to 1 when not stated.  If the document shows "Qty" or "Quantity" column, use that value.
3. **unit_price** — Price *per single unit* as a decimal number (no currency symbols, no commas).
   - If only a "Total Price" / "Extended Price" / "Amount" is given alongside a quantity > 1, compute unit_price = total / quantity.
   - If only a lump-sum total is shown with quantity = 1, use that total as unit_price.
   - If both Unit Price and Total Price columns exist, prefer the Unit Price column.

## How to Read Flattened Table Text
PDF text extraction often destroys table structure.  You may see column headers on one line followed by values on the next line(s).  Example:
```
#   Qty   Description            Unit Price   Total Price
1   5     Industrial Widget      $12.50       $62.50
2   10    Steel Bracket          $3.75        $37.50
```
Match values to their column headers by position.  The header row usually contains words like: #, Qty, Quantity, Description, Item, Unit Price, Price, Total, Amount, Extended, UOM.

Sometimes the text is fully linearised (one value per line):
```
#
Qty
Description of supplies or services
Unit Price
Total Price
1
1
Line item 1
$100.00
$100.00
```
In this case, group the values by repeating the header pattern (5 headers → every 5 values is one row).

## What to Exclude
- Summary/total rows (e.g. "Subtotal", "Tax", "Grand Total", "Net purchase not to exceed", "Gross purchase order", "Less supplier cost share")
- Shipping and handling charges that are not a purchasable item
- Discount lines (negative amounts that adjust other items)
- Header/footer boilerplate, page numbers, logos
- Terms and conditions text
- Approval signatures or workflow metadata

## Output Format
Return ONLY a JSON object:
```json
{"items": [{"description": "...", "quantity": 1, "unit_price": 0.00}, ...]}
```
If genuinely NO line items exist in the document, return: {"items": []}

CRITICAL: Do NOT invent items.  Extract ONLY what the document explicitly contains."""

# Image MIME types that GPT-4o vision can handle
IMAGE_TYPES = {"image/png", "image/jpeg", "image/jpg", "image/tiff", "image/gif", "image/webp", "image/bmp"}

# ── Deep-scan prompt (interpretive fallback) ────────────────────────
# Used when the standard extraction finds 0 items.  Much more liberal
# about what qualifies — anything with a dollar amount is a candidate.
DEEP_SCAN_PROMPT = """\
You are an expert document analyst performing a DEEP SCAN for potential \
line items that a standard extraction missed.

## Context
A previous extraction pass found ZERO line items in this document using \
strict table-parsing rules.  Your job is to look MORE BROADLY and \
INTERPRET the document content to find anything that could reasonably \
be treated as a purchasable line item.

## What to Look For (be generous)
- Any mention of a product, service, material, fee, charge, cost, \
expense, deliverable, or work item — especially if a dollar amount \
appears nearby.
- Dollar amounts anywhere in the document: "$1,200.00", "USD 500", \
"cost: 750.00", "price of $42", "fee: $300/hr", etc.
- Descriptions of work or services even without a formal table — \
e.g. "Consulting services — $5,000" or "Phase 1 design: $12,500".
- Lump-sum project costs, milestone payments, flat fees.
- Items described in paragraphs, bullet lists, or free-form text — \
not just tables.
- If a total project cost is given with a breakdown (percentages, \
phases, tasks), extract each component as a separate line item.
- If ONLY a single total cost is mentioned for the entire document \
with no breakdown, create ONE line item using the document subject \
as the description.

## Extraction Rules
1. **description** — Capture the most descriptive text near the amount.\
  Use the document's own wording.  Include part numbers, phase names, \
or reference numbers if present.
2. **quantity** — Default to 1 unless an explicit count is stated.
3. **unit_price** — The dollar amount as a decimal (no symbols/commas).\
  If a total is given with quantity > 1, compute unit_price = total / qty.

## What to Still Exclude
- Tax lines, shipping-only charges, discount adjustments
- Page numbers, dates, document IDs used as reference only
- Amounts that are clearly NOT prices (e.g. "Account #12345", \
"PO Number 50000")

## Output Format
Return ONLY a JSON object:
```json
{"items": [{"description": "...", "quantity": 1, "unit_price": 0.00}, ...]}
```
If after thorough analysis you are confident there are truly NO \
purchasable items or dollar amounts in the document, return: \
{"items": []}

Be thorough.  It is better to return a potential item that the user \
can remove than to miss a real one."""


def calculate_cost(prompt_tokens: int, completion_tokens: int) -> float:
    """Calculate estimated cost in USD from token counts."""
    input_cost = (prompt_tokens / 1_000_000) * COST_PER_1M_INPUT_TOKENS
    output_cost = (completion_tokens / 1_000_000) * COST_PER_1M_OUTPUT_TOKENS
    return round(input_cost + output_cost, 6)


def check_budget():
    """Check if AI spending is within budget.

    Returns:
        tuple: (within_budget, total_spent, budget_limit)
    """
    budget_limit = float(os.environ.get("AZURE_OPENAI_BUDGET_LIMIT", "100.00"))
    try:
        from models import AIUsageLog, db
        total_spent = db.session.query(
            db.func.coalesce(db.func.sum(AIUsageLog.estimated_cost_usd), 0)
        ).scalar()
        return (float(total_spent) < budget_limit, float(total_spent), budget_limit)
    except Exception:
        # AIUsageLog table may not exist yet — allow usage
        return (True, 0.0, budget_limit)


def _get_client():
    """Create Azure OpenAI client."""
    from openai import AzureOpenAI

    endpoint = os.environ.get("AZURE_OPENAI_ENDPOINT", "")
    key = os.environ.get("AZURE_OPENAI_KEY", "")

    if not endpoint or not key:
        raise ValueError("Azure OpenAI not configured. Set AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_KEY.")

    return AzureOpenAI(
        azure_endpoint=endpoint,
        api_key=key,
        api_version="2024-12-01-preview",
    )


def _extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF preserving table layout where possible.

    Strategy:
      1. pdfplumber with layout=True (best for tables)
      2. PyPDF2 fallback (if pdfplumber unavailable or fails)
    Returns empty string if the PDF is scanned/image-only.
    """
    # --- Strategy 1: pdfplumber (layout-aware) ---
    try:
        import pdfplumber

        text_parts = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                # layout=True preserves column alignment so the model can
                # parse tabular data by positional column matching.
                page_text = page.extract_text(layout=True)
                if page_text and page_text.strip():
                    text_parts.append(page_text)
        if text_parts:
            combined = "\n".join(text_parts)
            # Strip excessive leading whitespace on each line but keep relative
            # indentation by removing only truly blank prefix columns.
            lines = combined.split("\n")
            if lines:
                min_indent = min(
                    (len(l) - len(l.lstrip()) for l in lines if l.strip()),
                    default=0,
                )
                if min_indent > 0:
                    lines = [l[min_indent:] if len(l) > min_indent else l for l in lines]
            return "\n".join(lines)
    except ImportError:
        pass  # pdfplumber not installed — fall through to PyPDF2
    except Exception as e:
        logger.warning("pdfplumber extraction failed, falling back to PyPDF2: %s", e)

    # --- Strategy 2: PyPDF2 ---
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(file_path)
        text_parts = []
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
        return "\n".join(text_parts)
    except Exception as e:
        logger.warning("PDF text extraction failed: %s", e)
        return ""


def _extract_text_from_docx(file_path: str) -> str:
    """Extract text from Word document."""
    try:
        from docx import Document

        doc = Document(file_path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        logger.warning("DOCX text extraction failed: %s", e)
        return ""


def _extract_text_from_excel(file_path: str) -> str:
    """Extract text from Excel file as CSV-like text."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        text_parts = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    text_parts.append(" | ".join(cells))
        wb.close()
        return "\n".join(text_parts)
    except Exception as e:
        logger.warning("Excel text extraction failed: %s", e)
        return ""


def _extract_text_from_csv(file_path: str) -> str:
    """Extract text from CSV file."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            rows = []
            for i, row in enumerate(reader):
                if i >= 500:  # Cap at 500 rows
                    break
                rows.append(" | ".join(row))
        return "\n".join(rows)
    except Exception as e:
        logger.warning("CSV text extraction failed: %s", e)
        return ""


def _file_to_base64(file_path: str) -> str:
    """Read file and return base64-encoded string."""
    with open(file_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def _validate_items(raw_items: list) -> list:
    """Validate and sanitize extracted items."""
    validated = []
    for item in raw_items[:MAX_ITEMS]:
        if not isinstance(item, dict):
            continue
        desc = str(item.get("description", "")).strip()
        if not desc:
            continue
        try:
            qty = max(1, int(float(item.get("quantity", 1))))
        except (ValueError, TypeError):
            qty = 1
        try:
            price = max(0.0, round(float(item.get("unit_price", 0)), 2))
        except (ValueError, TypeError):
            price = 0.0
        validated.append({
            "description": desc[:500],  # Cap description length
            "quantity": qty,
            "unit_price": price,
        })
    return validated


def _call_gpt(client, deployment: str, messages: list) -> dict:
    """Send messages to GPT-4o and return parsed result dict.

    Returns dict with keys: items (list), prompt_tokens, completion_tokens,
    total_tokens, warning (str|None).
    """
    result = {"items": [], "prompt_tokens": 0, "completion_tokens": 0,
              "total_tokens": 0, "warning": None}

    try:
        response = client.chat.completions.create(
            model=deployment,
            messages=messages,
            response_format={"type": "json_object"},
            temperature=0.1,
            timeout=60,
        )

        result["prompt_tokens"] = response.usage.prompt_tokens
        result["completion_tokens"] = response.usage.completion_tokens
        result["total_tokens"] = response.usage.total_tokens

        cost = calculate_cost(response.usage.prompt_tokens,
                              response.usage.completion_tokens)
        logger.info(
            "AI extraction: %d prompt + %d completion = %d total tokens, "
            "est. cost $%.4f",
            response.usage.prompt_tokens,
            response.usage.completion_tokens,
            response.usage.total_tokens,
            cost,
        )

        content = response.choices[0].message.content
        parsed = json.loads(content)

        raw_items = parsed.get("items", [])
        if isinstance(raw_items, list):
            result["items"] = _validate_items(raw_items)
        else:
            result["warning"] = "AI returned unexpected format."

    except json.JSONDecodeError:
        logger.warning("Failed to parse AI response as JSON")
        result["warning"] = "AI returned invalid response format."
    except Exception as e:
        error_type = type(e).__name__
        logger.error("AI extraction failed (%s): %s", error_type, e)
        if "timeout" in error_type.lower() or "Timeout" in str(e):
            result["warning"] = "AI processing timed out. Try a smaller document."
        elif "RateLimit" in error_type:
            result["warning"] = (
                "AI service rate limit reached. Please try again in a moment."
            )
        else:
            result["warning"] = (
                "AI extraction encountered an error. Please try again."
            )

    return result


def _build_text_messages(document_text: str) -> list:
    """Build GPT messages for text-based extraction."""
    if len(document_text) > 80_000:
        document_text = (
            document_text[:80_000] + "\n\n[... document truncated ...]"
        )
    return [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": (
            "Extract all line items from this document.  "
            "The text below was extracted from a PDF or office document; "
            "table columns may be separated by whitespace rather than "
            "delimiters.\n\n" + document_text
        )},
    ]


def _pdf_page_to_base64(file_path: str) -> str:
    """Render the first page of a PDF to a PNG image, base64-encoded."""
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(file_path)
        page = doc[0]
        pix = page.get_pixmap(dpi=200)
        img_bytes = pix.tobytes("png")
        doc.close()
        return base64.b64encode(img_bytes).decode("utf-8")
    except Exception as e:
        logger.warning("PDF-to-image via fitz failed: %s", e)
        return _file_to_base64(file_path)


def _build_vision_messages(file_path: str, mime_lower: str,
                           system_prompt: str = None) -> list:
    """Build GPT messages for vision-based (image) extraction."""
    prompt = system_prompt or SYSTEM_PROMPT
    if mime_lower == "application/pdf":
        b64_data = _pdf_page_to_base64(file_path)
        img_mime = "image/png"
    else:
        b64_data = _file_to_base64(file_path)
        img_mime = (
            mime_lower if mime_lower in IMAGE_TYPES
            else "image/png"
        )
    return [
        {"role": "system", "content": prompt},
        {"role": "user", "content": [
            {"type": "text", "text": (
                "Extract all line items from this document image.  "
                "Look carefully for tables, numbered lists, or any "
                "rows containing item descriptions with quantities "
                "and prices."
            )},
            {"type": "image_url", "image_url": {
                "url": f"data:{img_mime};base64,{b64_data}",
            }},
        ]},
    ]


def _build_deep_scan_text_messages(document_text: str) -> list:
    """Build GPT messages for deep-scan (interpretive) extraction."""
    if len(document_text) > 80_000:
        document_text = (
            document_text[:80_000]
            + "\n\n[... document truncated ...]"
        )
    return [
        {"role": "system", "content": DEEP_SCAN_PROMPT},
        {"role": "user", "content": (
            "A standard line-item extraction found NOTHING in this "
            "document.  Please perform a deep scan — look for ANY "
            "mention of costs, fees, services, products, or dollar "
            "amounts that could be treated as line items.  Be "
            "thorough and interpret the content broadly.\n\n"
            + document_text
        )},
    ]


def extract_line_items(file_path: str, mime_type: str) -> dict:
    """Extract line items from a document using Azure OpenAI GPT-4o.

    Args:
        file_path: Path to the uploaded file
        mime_type: MIME type of the file

    Returns:
        dict with keys:
            items: list of {"description": str, "quantity": int, "unit_price": float}
            prompt_tokens: int
            completion_tokens: int
            total_tokens: int
            warning: str or None
    """
    deployment = os.environ.get("AZURE_OPENAI_DEPLOYMENT", "gpt-4o")
    result = {"items": [], "prompt_tokens": 0, "completion_tokens": 0,
              "total_tokens": 0, "warning": None}

    try:
        client = _get_client()
    except ValueError as e:
        result["warning"] = str(e)
        return result

    use_vision = False
    is_pdf = False
    document_text = ""

    # Determine extraction strategy based on MIME type
    mime_lower = (mime_type or "").lower()

    if mime_lower in IMAGE_TYPES:
        use_vision = True
    elif mime_lower == "application/pdf":
        is_pdf = True
        document_text = _extract_text_from_pdf(file_path)
        # If PDF has minimal text, treat as scanned image
        if len(document_text.strip()) < 50:
            use_vision = True
    elif mime_lower in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ):
        document_text = _extract_text_from_docx(file_path)
    elif mime_lower in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        document_text = _extract_text_from_excel(file_path)
    elif mime_lower == "text/csv":
        document_text = _extract_text_from_csv(file_path)
    elif mime_lower in ("text/plain", "text/markdown"):
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            document_text = f.read(100_000)
    else:
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                document_text = f.read(100_000)
        except Exception:
            result["warning"] = f"Unsupported file type: {mime_type}"
            return result

    if not use_vision and not document_text.strip():
        result["warning"] = (
            "No text content could be extracted from this document."
        )
        return result

    # --- Primary extraction attempt ---
    if use_vision:
        messages = _build_vision_messages(file_path, mime_lower)
    else:
        messages = _build_text_messages(document_text)

    result = _call_gpt(client, deployment, messages)

    # --- Vision fallback for PDFs ---
    # If text extraction returned content but GPT found 0 items, the table
    # structure may have been lost during text extraction.  Retry with
    # GPT-4o vision which can see the original layout.
    if is_pdf and not use_vision and len(result["items"]) == 0:
        logger.info(
            "Text extraction yielded 0 items for PDF — retrying with vision"
        )
        vision_messages = _build_vision_messages(file_path, "application/pdf")
        vision_result = _call_gpt(client, deployment, vision_messages)

        # Accumulate token usage from both attempts
        vision_result["prompt_tokens"] += result["prompt_tokens"]
        vision_result["completion_tokens"] += result["completion_tokens"]
        vision_result["total_tokens"] += result["total_tokens"]

        if vision_result["items"]:
            return vision_result
        # Vision also found nothing — accumulate tokens and fall through
        # to the deep-scan retry below.
        result["prompt_tokens"] = vision_result["prompt_tokens"]
        result["completion_tokens"] = vision_result["completion_tokens"]
        result["total_tokens"] = vision_result["total_tokens"]

    # --- Deep-scan fallback (interpretive) ---
    # All previous attempts (text + vision) found 0 items.  Retry with a
    # much more liberal prompt that interprets dollar amounts, free-form
    # descriptions, and non-tabular content as potential line items.
    if len(result["items"]) == 0 and not result.get("warning"):
        logger.info(
            "Standard extraction yielded 0 items — "
            "retrying with deep-scan prompt"
        )
        if use_vision or is_pdf:
            deep_msgs = _build_vision_messages(
                file_path,
                mime_lower if use_vision else "application/pdf",
                system_prompt=DEEP_SCAN_PROMPT,
            )
        else:
            deep_msgs = _build_deep_scan_text_messages(document_text)

        deep_result = _call_gpt(client, deployment, deep_msgs)

        # Accumulate token usage across all attempts
        deep_result["prompt_tokens"] += result["prompt_tokens"]
        deep_result["completion_tokens"] += result["completion_tokens"]
        deep_result["total_tokens"] += result["total_tokens"]

        if deep_result["items"]:
            deep_result["warning"] = (
                "Items found via deep interpretation — "
                "please review carefully before importing."
            )
            return deep_result

        # Truly nothing found
        result["prompt_tokens"] = deep_result["prompt_tokens"]
        result["completion_tokens"] = deep_result["completion_tokens"]
        result["total_tokens"] = deep_result["total_tokens"]
        result["warning"] = (
            "No line items detected after thorough analysis."
        )

    return result
